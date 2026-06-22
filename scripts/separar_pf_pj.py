#!/usr/bin/env python3
"""
Separador PF/PJ — Yago Modas
Lê Movimentação Diária + Predatados, classifica cada lançamento como PJ ou PF,
e gera uma planilha de conferência colorida para revisão antes de lançar no Controle.

Uso:
  python3 scripts/separar_pf_pj.py --mes 6 --ano 2026
  python3 scripts/separar_pf_pj.py --mes 6 --ano 2026 --listar-abas

Resultado: financeiro/conferencias/conferencia_JUN_2026.xlsx
"""

import argparse
import os
import re
import unicodedata
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import xlrd

# ──────────────────────────────────────────────────────────
# CAMINHOS
# ──────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DADOS = os.path.join(BASE, "financeiro", "dados-brutos")
SAIDA = os.path.join(BASE, "financeiro", "conferencias")

ARQ_MOV  = os.path.join(DADOS, "Movimentação Diária F1-1.xlsx")
ARQ_PRED = os.path.join(DADOS, "PREDATADOS F1-1.xls")

# ──────────────────────────────────────────────────────────
# TABELAS DE REFERÊNCIA
# ──────────────────────────────────────────────────────────
MESES_ABREV = {
    1:"Jan", 2:"Fev", 3:"Mar", 4:"Abr", 5:"Mai",
    6:"Jun", 7:"Jul", 8:"Ago", 9:"Set", 10:"Out", 11:"Nov", 12:"Dez"
}
MESES_NOME = {
    1:"JANEIRO", 2:"FEVEREIRO", 3:"MARÇO", 4:"ABRIL", 5:"MAIO",
    6:"JUNHO", 7:"JULHO", 8:"AGOSTO", 9:"SETEMBRO",
    10:"OUTUBRO", 11:"NOVEMBRO", 12:"DEZEMBRO"
}

# ──────────────────────────────────────────────────────────
# REGRAS DE CLASSIFICAÇÃO — ENCARGOS (PJ, col D da Movimentação)
# Descrição → (Classificação, Plano de Contas)
# ──────────────────────────────────────────────────────────
REGRAS_ENCARGOS = [
    # Funcionários
    (r"EMESCAM\s*NILTA|FGTS\s*NILTA|SEGURO\s*NILTA|V\.\s*NILTA",
     "DESPESA FUNCIONÁRIOS", "Nilta"),
    (r"\bFGTS\b", "IMPOSTOS", "FGTS"),
    (r"\bNILTA\b", "DESPESA FUNCIONÁRIOS", "Nilta"),           # flag ambígua — ver AMBIGUOS
    (r"CLAUDIA|CLÁUDIA", "DESPESA FUNCIONÁRIOS", "Cláudia"),
    (r"DAYANA|PGTO\s*DAYANA", "DESPESA FUNCIONÁRIOS", "Dayana"),
    (r"\bPEDRO\b", "DESPESA FUNCIONÁRIOS", "Pedro"),
    (r"EDYANDRA", "DESPESA FUNCIONÁRIOS", "Edyandra"),
    # Contador
    (r"JEAN\s*CONTABILIDADE|JEAN\s*CONTADOR|JEAN\b|CONTADOR\b|CONTABILIDADE\b",
     "DESPESAS OPERACIONAIS", "Contabilidade"),
    # Impostos
    (r"DARF|SIMPLES\s*NACIONAL|PGDAS|DAS\b", "IMPOSTOS", "DARF Simples"),
    (r"\bIPTU\b", "CUSTO COM INFRAESTRUTURA", "IPTU"),
    # Infraestrutura
    (r"ALUGUEL|ALUG\b", "CUSTO COM INFRAESTRUTURA", "Aluguel"),
    (r"GARAGEM", "CUSTO COM INFRAESTRUTURA", "Aluguel"),
    (r"ENERGIA|ELETRIC|CELESC|ESCELSA|\bEDP\b", "CUSTO COM INFRAESTRUTURA", "Energia"),
    (r"\bAGUA\b|SANEAMENTO|CESAN\b", "CUSTO COM INFRAESTRUTURA", "Água"),
    (r"INTERNET|BANDA\s*LARGA|VIVO\b|CLARO\b|\bTIM\b|\bOI\b",
     "CUSTO COM INFRAESTRUTURA", "Internet / Telefone"),
    # Carro empresa
    (r"\bMASTER\b|\bDUCATO\b", "CUSTO COM CARRO", "Despesas extras"),
    (r"RASTREADOR\b|TRACKER.*LJ|SEGURO.*CARRO|SEGURO.*MASTER|\bIPVA\b|\bDPVAT\b|DENATRAN|SENATRAN",
     "CUSTO COM CARRO", "Documentação"),
    (r"GASOLINA.*MASTER|GASOLINA.*DUCATO|GASOLINA.*LJ|GASOLINA.*DIESEL",
     "CUSTO COM CARRO", "Combustível"),
    # Financeiro
    (r"PRONAPE|IOF\b|JUROS\b|MULTA\s*ATRASO|ENCARGO\s*FINANC",
     "DESPESA FINANCEIRA", "Transferências bancárias"),
    (r"CONSÓRCIO|CONSORCIO", "INVESTIMENTO", "CONSÓRCIO"),
    # Operacional
    (r"PRODUTOS?\s*LIMPEZA|MATERIAL\s*LIMPEZA|MATERIAL\s*ESCRITÓRIO|PAPELARIA",
     "DESPESAS OPERACIONAIS", "Material de escritório / Limpeza"),
    (r"SISTEMA|SOFTWARE|APLICATIVO|PLATAFORMA",
     "DESPESAS OPERACIONAIS", "Software / Sistemas"),
    (r"CARTÓRIO|JUNTA\s*COMERCIAL|REGISTRO\b",
     "DESPESAS OPERACIONAIS", "Taxas e registros"),
    (r"ALMOÇO|REFEIÇÃO|ALIMENTAÇÃO", "OUTRA DESPESA", "DESPESA EXTRA"),
    (r"DESPESA|DESP\b|DESP\.", "OUTRA DESPESA", "DESPESA EXTRA"),
]

# PREDATADOS — keywords no campo FORNECEDOR → PF
REGRAS_PRED_PF = [
    r"CASA\b", r"KITINETE", r"JÂNIO|JANIO", r"\bIGOR\b", r"PRONAPE",
    r"FARMÁCIA|FARMACIA", r"MÉDICO|MEDICO", r"DENTISTA",
    r"SUPERMERCADO(?!.*LJ)", r"MERCADO(?!.*LJ)",
    r"GASOLINA.*TRACKER", r"GASOLINA.*PESSOAL",
    r"TRACKER(?!.*LJ|.*MASTER)",
    r"DISBAT.*MOTO",
    r"\bDEVOLUÇÃO\b",
]

# PREDATADOS — keywords no FORNECEDOR → PJ (fornecedores de mercadoria / empresa)
REGRAS_PRED_PJ = [
    r"P/LJ\b|P/ LJ|PLJ\b|PARA\s*LJ\b",
    r"\bMASTER\b|\bDIESEL\b|\bDUCATO\b",
    # fornecedores de mercadoria
    r"ADOMES", r"YANG\s*STYLE", r"LEPPER", r"\bV2\b", r"BIOGAS",
    r"MARJU", r"\bMTZ\b", r"URB\s*CONF", r"VEST\s*SURF",
    r"\bB\.MAR\b", r"BOKKER", r"\bALK\b", r"FRIGELAR",
    r"GRUPO\s*ELIAN", r"RILDENIS", r"\bATUAL\b", r"PISALE",
    r"R\.\s*MUNDU", r"AT\.SP", r"DANUBIO|DANUBIO",
    r"ACESSORIO\s*AUTOMOTIVO",
]

# Descrições/fornecedores que devem ser flagados para revisão manual
AMBIGUOS_DESCRICAO = [
    r"\bNILTA\b",          # pode ser PJ (funcionária loja) ou PF
    r"JEAN\b|CONTADOR\b",  # possível duplicidade PJ+PF
]

# ──────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────

def normalizar(s):
    s = s.strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)

def fuzzy_aba_predatados(wb_xls, mes, ano):
    """Localiza aba do mês/ano no PREDATADOS com normalização flexível."""
    mes_norm = normalizar(MESES_NOME[mes])
    ano_str  = str(ano)
    ano_2d   = ano_str[-2:]

    melhores = []
    for nome in wb_xls.sheet_names():
        nome_n = normalizar(nome).replace(" ", "")
        if ano_str not in nome_n and ano_2d not in nome_n:
            continue
        # Testar prefixos do mês: 3 chars em diante
        for pref_len in range(len(mes_norm), 2, -1):
            pref = mes_norm[:pref_len]
            if pref in nome_n:
                melhores.append((pref_len, nome))
                break
    if not melhores:
        return None
    melhores.sort(reverse=True)
    return melhores[0][1]

def classificar_encargos(descricao):
    desc = descricao.upper() if descricao else ""
    for padrao, classif, plano in REGRAS_ENCARGOS:
        if re.search(padrao, desc):
            return classif, plano
    return "DESPESAS OPERACIONAIS", "Outras despesas operacionais"

def checar_ambiguo(texto):
    if not texto:
        return False
    for p in AMBIGUOS_DESCRICAO:
        if re.search(p, texto.upper()):
            return True
    return False

def classificar_pred(fornecedor, ncheque=""):
    forn = fornecedor.upper().strip() if fornecedor else ""
    chq  = (ncheque or "").upper().strip()

    # N°CHEQUE explícito
    if chq == "CT PJ":
        return "PJ", "ALTO", "cartão PJ (CT PJ)"
    if chq == "CT":
        return "PF", "ALTO", "cartão pessoal (CT)"

    # Flag de ambiguidade
    if checar_ambiguo(forn):
        for p in AMBIGUOS_DESCRICAO:
            if re.search(p, forn):
                return "?", "BAIXO", f"AMBÍGUO — revisar: {re.search(p, forn).group()}"

    # Regras PF
    for kw in REGRAS_PRED_PF:
        if re.search(kw, forn, re.IGNORECASE):
            return "PF", "ALTO", f"keyword PF → {kw}"

    # Regras PJ
    for kw in REGRAS_PRED_PJ:
        if re.search(kw, forn, re.IGNORECASE):
            return "PJ", "ALTO", f"fornecedor PJ → {re.search(kw, forn, re.IGNORECASE).group()}"

    return "PJ", "MÉDIO", "sem regra → padrão PJ (conferir)"

def fmt_brl(v):
    if v is None:
        return ""
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ──────────────────────────────────────────────────────────
# LEITURA — MOVIMENTAÇÃO
# ──────────────────────────────────────────────────────────

def ler_movimentacao(mes, ano):
    wb = openpyxl.load_workbook(ARQ_MOV, read_only=True, data_only=True)
    ws = wb["2021"]
    lancamentos = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        data = row[0]
        if not isinstance(data, datetime):
            continue
        if data.month != mes or data.year != ano:
            continue

        descricao = str(row[1]).strip() if row[1] else ""
        compras   = row[2]   # PJ → CUSTO COM PRODUTOS
        encargos  = row[3]   # PJ → classificar por descrição
        desp_pf   = row[4]   # PF
        rec_pf    = row[5]   # PF
        recebt    = row[6]   # PJ → RECEITA COM PRODUTOS / Venda RUA
        loja      = row[7]   # PJ → RECEITA COM PRODUTOS / Loja

        if compras:
            lancamentos.append({
                "fonte": "MOV", "data": data, "descricao": descricao,
                "valor": compras, "tipo_mov": "SAÍDA",
                "destino": "PJ", "confianca": "ALTO",
                "classif": "CUSTO COM PRODUTOS", "plano": "Compra de Mercadoria",
                "forma": "Boleto", "banco": "Bradesco",
                "centro": "Loja", "cliente": "Loja Yago Modas",
                "flag": "⚠️ NILTA" if checar_ambiguo(descricao) else "",
            })
        if encargos:
            classif, plano = classificar_encargos(descricao)
            lancamentos.append({
                "fonte": "MOV", "data": data, "descricao": descricao,
                "valor": encargos, "tipo_mov": "SAÍDA",
                "destino": "PJ", "confianca": "ALTO",
                "classif": classif, "plano": plano,
                "forma": "Débito", "banco": "Bradesco",
                "centro": "Loja", "cliente": "Loja Yago Modas",
                "flag": "⚠️ AMBÍGUO — revisar" if checar_ambiguo(descricao) else "",
            })
        if desp_pf:
            lancamentos.append({
                "fonte": "MOV", "data": data, "descricao": descricao,
                "valor": desp_pf, "tipo_mov": "SAÍDA",
                "destino": "PF", "confianca": "ALTO",
                "classif": "DESPESA PESSOAL", "plano": "—",
                "forma": "Débito", "banco": "Bradesco",
                "centro": "—", "cliente": "—",
                "flag": "⚠️ AMBÍGUO — revisar" if checar_ambiguo(descricao) else "",
            })
        if rec_pf:
            lancamentos.append({
                "fonte": "MOV", "data": data, "descricao": descricao,
                "valor": rec_pf, "tipo_mov": "ENTRADA",
                "destino": "PF", "confianca": "ALTO",
                "classif": "RECEITA PESSOAL", "plano": "—",
                "forma": "—", "banco": "Bradesco",
                "centro": "—", "cliente": "—",
                "flag": "",
            })
        if recebt:
            lancamentos.append({
                "fonte": "MOV", "data": data, "descricao": descricao,
                "valor": recebt, "tipo_mov": "ENTRADA",
                "destino": "PJ", "confianca": "ALTO",
                "classif": "RECEITA COM PRODUTOS", "plano": "Venda de Mercadoria - RUA",
                "forma": "Dinheiro", "banco": "Bradesco",
                "centro": "Vendas Loja Externas", "cliente": "Loja Yago Modas",
                "flag": "",
            })
        if loja:
            lancamentos.append({
                "fonte": "MOV", "data": data, "descricao": descricao,
                "valor": loja, "tipo_mov": "ENTRADA",
                "destino": "PJ", "confianca": "ALTO",
                "classif": "RECEITA COM PRODUTOS", "plano": "Venda de Mercadoria Yago Modas",
                "forma": "Dinheiro", "banco": "Bradesco",
                "centro": "Vendas Loja Interna", "cliente": "Loja Yago Modas",
                "flag": "",
            })
    wb.close()
    return lancamentos

# ──────────────────────────────────────────────────────────
# LEITURA — PREDATADOS
# ──────────────────────────────────────────────────────────

def ler_predatados(mes, ano):
    wb = xlrd.open_workbook(ARQ_PRED)
    nome_aba = fuzzy_aba_predatados(wb, mes, ano)
    if not nome_aba:
        print(f"  [AVISO] Aba de predatados para {MESES_NOME[mes]}/{ano} não encontrada.")
        return [], None

    ws = wb.sheet_by_name(nome_aba)
    lancamentos = []

    for i in range(2, ws.nrows):  # pular linha 0 (título) e 1 (header)
        row = ws.row_values(i)
        if not row[0]:  # sem data de vencimento
            continue

        # Data de vencimento (serial Excel)
        try:
            dt = xlrd.xldate_as_datetime(row[0], wb.datemode)
        except Exception:
            continue

        fornecedor  = str(row[1]).strip() if row[1] else ""
        parcelas    = str(row[2]).strip() if row[2] else ""
        especif     = str(row[3]).strip() if row[3] else ""
        ncheque     = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        conta       = str(row[7]).strip() if len(row) > 7 and row[7] else ""

        try:
            valor = float(row[8]) if row[8] else None
        except (ValueError, TypeError):
            valor = None

        if not valor:
            continue

        baixado = str(row[9]).strip() if len(row) > 9 and row[9] else ""
        destino, confianca, motivo = classificar_pred(fornecedor, ncheque)

        lancamentos.append({
            "fonte": f"PRED/{nome_aba}", "data": dt,
            "descricao": fornecedor + (f" [{parcelas}]" if parcelas else ""),
            "valor": valor, "tipo_mov": "SAÍDA",
            "destino": destino, "confianca": confianca,
            "classif": "CUSTO COM PRODUTOS" if destino == "PJ" else "DESPESA PESSOAL",
            "plano": "Compra de Mercadoria" if destino == "PJ" else "—",
            "forma": "Boleto",
            "banco": conta if conta else "Bradesco",
            "centro": "Loja" if destino == "PJ" else "—",
            "cliente": "Loja Yago Modas" if destino == "PJ" else "—",
            "flag": f"⚠️ {motivo}" if (destino == "?" or confianca == "MÉDIO") else "",
            "_motivo_class": motivo,
            "_baixado": baixado,
        })

    return lancamentos, nome_aba

# ──────────────────────────────────────────────────────────
# GERAÇÃO — PLANILHA DE CONFERÊNCIA (Excel colorida)
# ──────────────────────────────────────────────────────────

# Paleta de cores
COR = {
    "verde_bg":   PatternFill("solid", fgColor="D6F4E8"),   # PJ alta confiança
    "azul_bg":    PatternFill("solid", fgColor="D6E8FF"),   # PF alta confiança
    "amarelo_bg": PatternFill("solid", fgColor="FFF3CD"),   # médio / sem regra
    "vermelho_bg":PatternFill("solid", fgColor="FFE0DC"),   # ambíguo / flag
    "cinza_bg":   PatternFill("solid", fgColor="F2F2F2"),   # cabeçalho
    "header_bg":  PatternFill("solid", fgColor="1E3A5F"),
}

def cell_fill(row_data):
    dest = row_data.get("destino", "")
    conf = row_data.get("confianca", "")
    flag = row_data.get("flag", "")
    if flag or dest == "?":
        return COR["vermelho_bg"]
    if conf == "MÉDIO":
        return COR["amarelo_bg"]
    if dest == "PJ":
        return COR["verde_bg"]
    if dest == "PF":
        return COR["azul_bg"]
    return None

def gerar_conferencia(lancamentos, mes, ano, aba_pred):
    os.makedirs(SAIDA, exist_ok=True)
    mes_str = MESES_ABREV[mes].upper()
    nome_arq = os.path.join(SAIDA, f"conferencia_{mes_str}_{ano}.xlsx")

    wb = openpyxl.Workbook()

    # ── Aba: CONFERÊNCIA (todos os lançamentos) ──
    ws = wb.active
    ws.title = "CONFERÊNCIA"

    colunas = [
        ("DATA",        12),
        ("FONTE",       18),
        ("DESCRIÇÃO",   38),
        ("VALOR",       14),
        ("TIPO",         8),
        ("DESTINO",      9),
        ("CONFIANÇA",   10),
        ("CLASSIFICAÇÃO", 26),
        ("PLANO DE CONTAS", 32),
        ("FORMA PGTO",  14),
        ("BANCO",       14),
        ("CENTRO",      22),
        ("CLIENTE",     22),
        ("FLAG",        40),
    ]

    hd_font  = Font(color="FFFFFF", bold=True, size=10)
    hd_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, (nome, larg) in enumerate(colunas, 1):
        c = ws.cell(row=1, column=ci, value=nome)
        c.font = hd_font
        c.fill = COR["header_bg"]
        c.alignment = hd_align
        ws.column_dimensions[c.column_letter].width = larg

    ws.row_dimensions[1].height = 28

    for ri, lc in enumerate(lancamentos, 2):
        fill = cell_fill(lc)
        vals = [
            lc["data"].strftime("%d/%m/%Y") if isinstance(lc["data"], datetime) else str(lc["data"]),
            lc["fonte"],
            lc["descricao"],
            lc["valor"],
            lc["tipo_mov"],
            lc["destino"],
            lc["confianca"],
            lc["classif"],
            lc["plano"],
            lc["forma"],
            lc["banco"],
            lc["centro"],
            lc["cliente"],
            lc.get("flag", ""),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            if fill:
                c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 14))
            if ci == 4:  # VALOR
                c.number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{len(lancamentos)+1}"

    # ── Aba: RESUMO ──
    ws2 = wb.create_sheet("RESUMO")

    pj_lc = [l for l in lancamentos if l["destino"] == "PJ"]
    pf_lc = [l for l in lancamentos if l["destino"] == "PF"]
    amb_lc = [l for l in lancamentos if l["destino"] == "?"]
    med_lc = [l for l in lancamentos if l.get("confianca") == "MÉDIO"]
    flag_lc = [l for l in lancamentos if l.get("flag")]

    pj_ent  = sum(l["valor"] for l in pj_lc if l["tipo_mov"] == "ENTRADA")
    pj_said = sum(l["valor"] for l in pj_lc if l["tipo_mov"] == "SAÍDA")
    pf_ent  = sum(l["valor"] for l in pf_lc if l["tipo_mov"] == "ENTRADA")
    pf_said = sum(l["valor"] for l in pf_lc if l["tipo_mov"] == "SAÍDA")

    linhas_resumo = [
        ("RESUMO — CONFERÊNCIA PF/PJ", f"{MESES_NOME[mes]} {ano}", "", ""),
        ("", "", "", ""),
        ("", "LANÇAMENTOS", "ENTRADAS", "SAÍDAS"),
        ("PJ (empresa)",  len(pj_lc),  fmt_brl(pj_ent),  fmt_brl(pj_said)),
        ("PF (pessoal)",  len(pf_lc),  fmt_brl(pf_ent),  fmt_brl(pf_said)),
        ("AMBÍGUO / revisar", len(amb_lc), "—", "—"),
        ("MÉDIO (sem regra)", len(med_lc), "—", "—"),
        ("", "", "", ""),
        ("FLAGS DE ATENÇÃO", "", "", ""),
    ]
    for lc in flag_lc:
        linhas_resumo.append((
            lc["data"].strftime("%d/%m") if isinstance(lc["data"], datetime) else "?",
            lc["descricao"][:40],
            fmt_brl(lc["valor"]),
            lc.get("flag", ""),
        ))

    for ri, row in enumerate(linhas_resumo, 1):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            if ri == 1 and ci == 1:
                c.font = Font(bold=True, size=13)
            elif ri in (3, 9):
                c.font = Font(bold=True)

    for ci, larg in [(1, 22), (2, 36), (3, 18), (4, 40)]:
        ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = larg

    # ── Aba: LEGENDA ──
    ws3 = wb.create_sheet("LEGENDA")
    legenda = [
        ("COR", "SIGNIFICADO"),
        ("Verde",    "PJ — alta confiança, pode lançar diretamente"),
        ("Azul",     "PF — alta confiança, lançar no Controle Pessoal"),
        ("Amarelo",  "PJ provável — sem regra explícita, revisar antes de lançar"),
        ("Vermelho", "AMBÍGUO ou flagado — revisão manual obrigatória"),
        ("", ""),
        ("DESTINO", ""),
        ("PJ",  "Lançar em 1__CONTROLE_PJ.xlsx → aba do mês"),
        ("PF",  "Lançar em 2__CONTROLE_PESSOAL.xlsx → aba do mês"),
        ("?",   "Não lançar automaticamente — confirmar com Yago"),
        ("", ""),
        ("CONFIANÇA", ""),
        ("ALTO",   "Regra explícita ou coluna já indica classificação"),
        ("MÉDIO",  "Padrão PJ aplicado — sem correspondência de regra"),
        ("BAIXO",  "Ambíguo, possível erro — revisar obrigatoriamente"),
    ]
    fills_legenda = {
        "Verde":    COR["verde_bg"],
        "Azul":     COR["azul_bg"],
        "Amarelo":  COR["amarelo_bg"],
        "Vermelho": COR["vermelho_bg"],
    }
    for ri, (k, v) in enumerate(legenda, 1):
        c1 = ws3.cell(row=ri, column=1, value=k)
        c2 = ws3.cell(row=ri, column=2, value=v)
        if k in fills_legenda:
            c1.fill = fills_legenda[k]
            c2.fill = fills_legenda[k]
        if k in ("COR", "DESTINO", "CONFIANÇA"):
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 60

    wb.save(nome_arq)
    return nome_arq

# ──────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Separador PF/PJ — Yago Modas")
    parser.add_argument("--mes",  type=int, default=datetime.now().month)
    parser.add_argument("--ano",  type=int, default=datetime.now().year)
    parser.add_argument("--listar-abas", action="store_true",
                        help="Lista as abas disponíveis no arquivo de Predatados e sai")
    args = parser.parse_args()

    if args.listar_abas:
        wb = xlrd.open_workbook(ARQ_PRED)
        print(f"\nAbas no PREDATADOS ({len(wb.sheet_names())} total):")
        for n in wb.sheet_names():
            print(f"  '{n}'")
        return

    mes, ano = args.mes, args.ano
    print(f"\n{'='*60}")
    print(f"  SEPARADOR PF/PJ — {MESES_NOME[mes]} {ano}")
    print(f"{'='*60}")

    print("\n[1/3] Lendo Movimentação Diária…")
    lc_mov = ler_movimentacao(mes, ano)
    print(f"      {len(lc_mov)} lançamentos encontrados ({MESES_NOME[mes]}/{ano})")

    print("\n[2/3] Lendo Predatados…")
    lc_pred, aba_pred = ler_predatados(mes, ano)
    if aba_pred:
        print(f"      Aba encontrada: '{aba_pred}' — {len(lc_pred)} lançamentos")
    else:
        print("      Nenhum predatado encontrado para este mês.")

    todos = lc_mov + lc_pred
    todos.sort(key=lambda x: x["data"])

    pj   = sum(1 for l in todos if l["destino"] == "PJ")
    pf   = sum(1 for l in todos if l["destino"] == "PF")
    amb  = sum(1 for l in todos if l["destino"] == "?")
    med  = sum(1 for l in todos if l.get("confianca") == "MÉDIO")
    flags = sum(1 for l in todos if l.get("flag"))

    print(f"\n  Total: {len(todos)} lançamentos")
    print(f"  ✅ PJ:      {pj}")
    print(f"  🔵 PF:      {pf}")
    print(f"  ⚠️  Ambíguo: {amb}")
    print(f"  🟡 Médio:   {med}")
    print(f"  🔴 Flags:   {flags}")

    print("\n[3/3] Gerando planilha de conferência…")
    arq = gerar_conferencia(todos, mes, ano, aba_pred)
    print(f"\n  ✅ Salvo em: {arq}")
    print("\n  Próximo passo: abrir o arquivo, revisar os itens em VERMELHO e AMARELO,")
    print("  confirmar a classificação e lançar nas planilhas de controle.")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
